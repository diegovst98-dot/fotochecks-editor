# Tests del Editor de Fotos DISECOD — el candado de calidad antes de publicar.
#
# Uso:
#   python tests\correr_tests.py            -> corre todo; codigo de salida 1 si algo falla
#   python tests\correr_tests.py --aprobar  -> regenera las salidas "doradas" (hacerlo SOLO
#                                              tras un cambio intencional ya revisado a ojo)
#
# publicar.py corre esto automaticamente: si falla, NO se publica.
#
# Que cubre:
#   1. Que todo el codigo compile.
#   2. Unidades: emparejado con Excel, revision de pedido, mensaje al cliente,
#      registro de lotes, hoja de aprobacion.
#   3. DORADOS: las 10 fotos reales de entrada/ procesadas en transparente y
#      fondo blanco + 2 firmas sinteticas, comparadas pixel a pixel contra las
#      salidas aprobadas en tests/doradas/ (que viven SOLO en esta PC, nunca
#      en GitHub: son fotos de clientes).
#   4. Humo de la interfaz (que la ventana arme con sus 3 pestañas).

import os
import shutil
import subprocess
import sys
import time
from pathlib import Path

AQUI = Path(__file__).resolve().parent
BASE = AQUI.parent
sys.path.insert(0, str(BASE / "codigo"))

DORADAS = AQUI / "doradas"
TMP = AQUI / "tmp"
APROBAR = "--aprobar" in sys.argv

fallas = []


def check(nombre, cond, detalle=""):
    if cond:
        print(f"  OK    {nombre}")
    else:
        print(f"  FALLA {nombre}  {detalle}")
        fallas.append(nombre)


def comparar_imagenes(ruta_nueva, ruta_dorada):
    # Igualdad practica: tolera ruido de decodificacion, NO cambios visibles.
    import numpy as np
    from PIL import Image
    a = np.asarray(Image.open(ruta_nueva).convert("RGBA")).astype(np.float32)
    b = np.asarray(Image.open(ruta_dorada).convert("RGBA")).astype(np.float32)
    if a.shape != b.shape:
        return False, f"tamano distinto {a.shape} vs {b.shape}"
    dif = np.abs(a - b).mean(axis=2)
    frac = float((dif > 8).mean())
    media = float(dif.mean())
    if frac > 0.002 or media > 0.5:
        return False, f"{frac*100:.2f}% de pixeles distintos (media {media:.2f})"
    return True, ""


def main():
    t0 = time.time()
    if TMP.exists():
        shutil.rmtree(TMP)
    TMP.mkdir(parents=True)
    DORADAS.mkdir(parents=True, exist_ok=True)

    # ---------- 1. el codigo compila ----------
    print("\n[1/6] Compilacion")
    for p in sorted((BASE / "codigo").glob("*.py")):
        r = subprocess.run([sys.executable, "-m", "py_compile", str(p)],
                           capture_output=True)
        check(f"compila {p.name}", r.returncode == 0,
              r.stderr.decode(errors="ignore")[:200])

    # test rapido de _alfa_apretado (sin modelo): contrato del borde firme
    r = subprocess.run([sys.executable, str(AQUI / "test_alfa.py")],
                       capture_output=True)
    check("test_alfa (borde apretado)", r.returncode == 0,
          r.stdout.decode(errors="ignore")[-200:])

    if fallas:  # sin compilacion no tiene sentido seguir
        return

    # ---------- 1b. imports compatibles con el .exe ----------
    # El exe congelado solo trae las piezas de Python empacadas al CONSTRUIRLO;
    # un import nuevo en codigo/ que no este en el paquete ROMPE el programa en
    # las PCs del equipo aunque aqui funcione (leccion v21: logging.handlers).
    # Limite conocido: 'from X import Y' solo audita X, no el submodulo Y.
    print("\n[1b] Imports compatibles con el exe")
    import ast
    import zipfile
    imports = set()
    propios = {p.stem for p in (BASE / "codigo").glob("*.py")}
    for p in (BASE / "codigo").glob("*.py"):
        arbol = ast.parse(p.read_text(encoding="utf-8"))
        for nodo in ast.walk(arbol):
            if isinstance(nodo, ast.Import):
                imports.update(a.name for a in nodo.names)
            elif isinstance(nodo, ast.ImportFrom) and nodo.module:
                imports.add(nodo.module)
    imports = {i for i in imports if i.split(".")[0] not in propios}
    internal = BASE / "dist" / "FotochecksEditor" / "_internal"
    exe = BASE / "dist" / "FotochecksEditor" / "FotochecksEditor.exe"
    if not exe.exists():
        check("exe disponible para auditar imports", False,
              "no hay dist/ en esta PC: no se puede garantizar compatibilidad")
    else:
        disponibles = set(sys.builtin_module_names)
        blz = internal / "base_library.zip"
        if blz.exists():
            for n in zipfile.ZipFile(blz).namelist():
                m = n.replace("/", ".").removesuffix(".pyc").removesuffix(".py")
                if m.endswith(".__init__"):
                    m = m[:-len(".__init__")]
                disponibles.add(m)
        try:
            from PyInstaller.archive.readers import CArchiveReader
            arch = CArchiveReader(str(exe))
            for nombre in list(arch.toc):
                if str(nombre).endswith(".pyz"):
                    disponibles.update(
                        arch.open_embedded_archive(nombre).toc.keys())
        except Exception as e:
            print(f"  (aviso: no se pudo leer el PYZ del exe: {e})")
        carpetas = set()
        for d in internal.iterdir():
            if d.is_dir():
                carpetas.add(d.name)
            elif d.suffix == ".pyd":
                disponibles.add(d.name.split(".")[0])
        for imp in sorted(imports):
            raiz = imp.split(".")[0]
            ok = (imp in disponibles or raiz in carpetas
                  or ("." not in imp and raiz in disponibles))
            check(f"el exe trae '{imp}'", ok,
                  "NO esta en el exe: quitarlo del codigo o reconstruir el exe")

    import editar_fotos as core
    from PIL import Image, ImageDraw, ImageFilter
    import numpy as np

    # ---------- 2. emparejado con Excel ----------
    print("\n[2/6] Emparejado con Excel")
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Codigo", "Nombre"])
    for fila in (("1001", "Juan Perez Gomez"), ("1002", "Maria Lopez Diaz"),
                 ("1003", "Maria Lopez Castro"), ("1004", "Ana Torres Vega"),
                 ("1005", "Jorge Quispe Mamani")):   # 1005 no tendra foto
        ws.append(fila)
    xls = TMP / "personal.xlsx"
    wb.save(xls)
    codigos = core.cargar_codigos(xls)
    check("lee 5 registros del Excel", len(codigos) == 5, f"leyo {len(codigos)}")
    c, e = core.emparejar("Juan Perez Gomez", codigos)
    check("match exacto", c == "1001" and e in ("exacto", "aproximado"), f"{c}/{e}")
    c, e = core.emparejar("ana torres", codigos)
    check("match aproximado (minusculas, 2 tokens)", c == "1004", f"{c}/{e}")
    c, e = core.emparejar("Maria Lopez", codigos)
    check("ambiguo detectado (2 Marias)", e == "ambiguo", f"{c}/{e}")
    c, e = core.emparejar("Pedro Inexistente", codigos)
    check("sin match detectado", not c or e not in ("exacto", "aproximado"), f"{c}/{e}")

    # typos de una persona unica -> auto (aproximado); ver docs/2026-06-17-matching-*
    wb2 = Workbook(); ws2 = wb2.active
    ws2.append(["Codigo", "Nombre"])
    for fila in (("2001", "Cielo Arroyo"), ("2002", "Cristhian Camargo"),
                 ("2003", "Mary Carmen Condori"), ("2004", "Ana Flores Vega"),
                 ("2005", "Ana Torres Vega")):
        ws2.append(fila)
    xls2 = TMP / "personal2.xlsx"; wb2.save(xls2)
    cods2 = core.cargar_codigos(xls2)
    c, e = core.emparejar("Cielo Aroyo", cods2)
    check("typo auto: Aroyo->Arroyo", c == "2001" and e == "aproximado", f"{c}/{e}")
    c, e = core.emparejar("Cristian Camargo", cods2)
    check("typo auto: Cristian->Cristhian", c == "2002" and e == "aproximado", f"{c}/{e}")
    c, e = core.emparejar("Mary CarmenCondori", cods2)
    check("typo auto: CarmenCondori->Carmen Condori", c == "2003" and e == "aproximado", f"{c}/{e}")
    c, e = core.emparejar("Ana Vega", cods2)
    check("seguridad: parcial compartido NO se auto-empareja", e == "ambiguo", f"{c}/{e}")

    # P2/P3: lectura multi-columna (codigo=DNI, detalle p/ homonimos) + alerta DNI
    wb3 = Workbook(); ws3 = wb3.active
    ws3.append(["EMPRESA", "TRABAJADOR", "CARGO", "DNI"])
    ws3.append(["ACME", "Ricardo Flores", "Supervisor", "73217258"])
    ws3.append(["ACME", "Ricardo Flores", "CEO", "15726725"])
    xls3 = TMP / "personal3.xlsx"; wb3.save(xls3)
    cods3 = core.cargar_codigos(xls3)
    check("P2 multi-col: codigo=DNI y trae detalle",
          len(cods3) == 2 and all(r.get("detalle") for r in cods3), str(cods3))
    check("P3 alerta DNI: detecta 7 digitos",
          bool(core.dni_sospechosos([{"codigo": "7877823", "nombre": "X"}])), "")

    # ---------- 3. revision de pedido + mensaje ----------
    print("\n[3/6] Revision de pedido")
    entrada = sorted(p for p in (BASE / "entrada").iterdir()
                     if p.suffix.lower() in (".jpg", ".jpeg", ".png"))
    check("hay fotos reales en entrada/ (esta PC es la que publica)",
          len(entrada) >= 6, f"hay {len(entrada)}")
    if len(entrada) >= 6:
        pedido = TMP / "pedido"
        pedido.mkdir()
        nombres = ["Juan Perez Gomez", "Maria Lopez Diaz", "Ana Torres Vega"]
        for foto, n in zip(entrada[:3], nombres):
            shutil.copy2(foto, pedido / (n + ".jpg"))
        img = Image.open(entrada[0]).filter(ImageFilter.GaussianBlur(8))
        img.save(pedido / "Maria Lopez Castro.jpg", quality=85)   # borrosa
        Image.open(entrada[1]).resize((120, 160)).save(pedido / "Visitante Extra.jpg")
        fotos_pedido = sorted(pedido.iterdir())
        rev = core.revisar_fotos(fotos_pedido, codigos)
        check("revision: total correcto", rev["total"] == 5, str(rev["total"]))
        problemas = {f["nombre"]: f["problemas"] for f in rev["con_problema"]}
        check("revision: detecta la borrosa",
              any("borrosa" in " ".join(v) for k, v in problemas.items()
                  if k.startswith("Maria Lopez Castro")), str(problemas))
        check("revision: detecta resolucion baja",
              any("resolucion" in " ".join(v) for v in problemas.values()), str(problemas))
        check("revision: detecta al que no esta en el Excel",
              any("Excel" in " ".join(v) for k, v in problemas.items()
                  if k.startswith("Visitante")), str(problemas))
        check("revision: detecta persona sin foto",
              any("1005" in s for s in rev["sin_foto"]), str(rev["sin_foto"]))
        msj = core.mensaje_para_cliente(rev)
        check("mensaje: pide reenvios", "REENVIAR" in msj, "")
        check("mensaje: lista a los sin foto", "FALTAN LAS FOTOS" in msj, "")
        rev_ok = core.revisar_fotos(fotos_pedido[:1])
        if not rev_ok["con_problema"]:
            check("mensaje: caso todo conforme",
                  "conforme" in core.mensaje_para_cliente(rev_ok), "")

        # por_confirmar + aplicar_resoluciones (Fase 2): un nombre ambiguo
        # (2 Marias) queda para elegir a mano; resolverlo limpia el mensaje.
        amb = pedido / "Maria Lopez.jpg"
        shutil.copy2(entrada[0], amb)
        rev2 = core.revisar_fotos([amb], codigos)
        check("por_confirmar: detecta el ambiguo",
              any(c["nombre"] == "Maria Lopez.jpg" and len(c["candidatos"]) == 2
                  for c in rev2["por_confirmar"]), str(rev2["por_confirmar"]))
        core.aplicar_resoluciones(rev2, {"Maria Lopez.jpg": "1002"})
        check("aplicar_resoluciones: limpia por_confirmar",
              not rev2["por_confirmar"], str(rev2["por_confirmar"]))
        check("aplicar_resoluciones: saca a la persona de sin_foto",
              not any(s.startswith("1002") for s in rev2["sin_foto"]), str(rev2["sin_foto"]))

        # P4: override de calidad ("esta foto va igual") limpia el aviso de calidad
        rev_q = {"total": 1, "fotos": [{"nombre": "x.jpg", "estado": None,
                 "candidatos": [], "problemas": ["se ve borrosa"]}],
                 "sin_foto": [], "por_confirmar": [],
                 "por_calidad": [{"nombre": "x.jpg", "problemas": ["se ve borrosa"]}],
                 "con_problema": [], "ok": 0}
        core.aplicar_resoluciones(rev_q, {}, {"x.jpg"})
        check("P4 calidad_ok: limpia el problema de calidad",
              not rev_q["con_problema"] and not rev_q["por_calidad"], str(rev_q))
        # P5: reporte CSV se genera
        rep = TMP / "reporte.csv"
        core.reporte_csv(rev2, rep)
        check("P5 reporte_csv genera archivo",
              rep.exists() and rep.stat().st_size > 0, str(rep))

        # N5: el saludo del mensaje nombra al cliente
        rev_n = {"total": 1, "fotos": [], "sin_foto": ["1009 - Falta Uno"],
                 "duplicados": [], "por_confirmar": [], "por_calidad": [],
                 "con_problema": [], "ok": 0, "con_excel": True}
        check("N5 saludo nombra al cliente",
              "ACME SAC" in core.mensaje_para_cliente(rev_n, "ACME SAC"), "")
        # N3: el aviso HEIC entra en la seccion REENVIAR del mensaje
        rev_h = {"total": 1, "sin_foto": [], "duplicados": [], "por_confirmar": [],
                 "por_calidad": [], "ok": 0, "con_excel": True,
                 "fotos": [{"nombre": "f.heic", "estado": "error", "candidatos": [],
                            "problemas": ["es foto de iPhone (HEIC); pidela exportada como JPG"]}]}
        rev_h["con_problema"] = [f for f in rev_h["fotos"] if f["problemas"]]
        msj_h = core.mensaje_para_cliente(rev_h)
        check("N3 HEIC va en REENVIAR", "REENVIAR" in msj_h and "HEIC" in msj_h, msj_h)

        # NUEVA PESTAÑA: plan_cardpresso + aplicar_cardpresso (renombrar a <DNI>)
        from pathlib import Path as _P
        cp_fotos = [_P(entrada[0]), _P(entrada[1])]   # fotos reales
        cp_dir = pedido / "_cp_in"; cp_dir.mkdir(exist_ok=True)
        f_juan = cp_dir / "Juan Perez Gomez.jpg"; shutil.copy2(entrada[0], f_juan)
        f_amb = cp_dir / "Maria Lopez.jpg"; shutil.copy2(entrada[1], f_amb)
        pl = core.plan_cardpresso([f_juan, f_amb], codigos)
        check("cardpresso: cruza exacto (Juan->1001)",
              any(p["nombre"] == "Juan Perez Gomez.jpg" and p["codigo"] == "1001"
                  for p in pl["plan"]), str(pl["plan"]))
        check("cardpresso: ambiguo va a por_confirmar",
              any(c["nombre"] == "Maria Lopez.jpg" for c in pl["por_confirmar"]),
              str(pl["por_confirmar"]))
        # resolver el ambiguo a mano y aplicar
        pl2 = core.plan_cardpresso([f_juan, f_amb], codigos, {"Maria Lopez.jpg": "1002"})
        cp_out = TMP / "cp_out"
        res = core.aplicar_cardpresso(pl2["plan"], cp_out, formato_unico=False)
        check("cardpresso: copia las 2 renombradas", res["copiadas"] == 2, str(res))
        check("cardpresso: archivo <DNI>.jpg existe",
              (cp_out / "1001.jpg").exists() and (cp_out / "1002.jpg").exists(),
              str(sorted(p.name for p in cp_out.iterdir())))
        check("cardpresso: genera _verificacion.csv",
              (cp_out / "_verificacion.csv").exists(), "")

    # ---------- 4. registro de lotes y hoja de aprobacion ----------
    print("\n[4/6] Lotes y hoja de aprobacion")
    base_real = core.BASE
    try:
        core.BASE = TMP  # que el test no toque el lotes.csv real
        core.registrar_lote("Cliente Test", 7, 5, TMP)
        core.registrar_lote("", 3, 0, TMP)
        csv = (TMP / "lotes.csv").read_text(encoding="utf-8-sig")
        check("lotes.csv: cabecera + 2 filas", csv.count("\n") >= 3, csv[:120])
        check("lotes.csv: registra el cliente", "Cliente Test;7;5" in csv, csv)
        c1 = core.carpeta_pedido("ACME S.A.C.")
        c2 = core.carpeta_pedido("ACME S.A.C.")
        check("carpeta de pedido: fecha+cliente y reutilizable",
              c1.exists() and c1 == c2 and c1.parent.name == "pedidos"
              and c1.name.endswith("ACME SAC"), str(c1))
        c3 = core.carpeta_pedido("")
        check("carpeta de pedido sin cliente", c3.name.endswith("pedido"), str(c3))
    finally:
        core.BASE = base_real
    if len(entrada) >= 3:
        pdf = core.hoja_aprobacion(entrada[:3], TMP / "ap.pdf", "Empresa Test",
                                   {"1001": "Juan Perez"})
        check("hoja de aprobacion: PDF generado", pdf.exists() and pdf.stat().st_size > 10000,
              f"{pdf.stat().st_size if pdf.exists() else 0} bytes")

    # recorte dudoso: dos mascaras casi iguales = NO dudoso; muy distintas = SI
    base = np.zeros((400, 400), np.uint8)
    base[100:300, 100:300] = 255
    casi = base.copy()
    casi[100:300, 100:205] = 255  # diferencia minima (<1.2%)
    check("recorte dudoso: mascaras iguales no se marcan",
          core.recorte_dudoso(base, casi) is False, "")
    muy = np.zeros((400, 400), np.uint8)
    muy[100:300, 180:380] = 255  # silueta corrida: gran desacuerdo
    check("recorte dudoso: mascaras muy distintas se marcan",
          core.recorte_dudoso(base, muy) is True, "")

    # ---------- 5. DORADOS: el pipeline completo contra salidas aprobadas ----------
    print("\n[5/6] Salidas doradas (esto carga el modelo, ~20s)")
    preset = core.cargar_preset()
    preset.update({"formato_salida": "PNG", "brillo": 1.32, "brillo_auto": False,
                   "piso_negro": 0, "ancho_px": 1067, "alto_px": 1031,
                   "color_auto": False, "saturacion_auto": False})
    session, fino = core.sesion_recorte(preset)
    check("modelo fino activo", fino, "cayo al clasico: ¿falta isnet o internet?")

    # dos firmas sinteticas deterministas
    firmas_dir = TMP / "firmas"
    firmas_dir.mkdir()
    import math
    f1 = Image.new("RGB", (1100, 700), (252, 252, 250))
    d = ImageDraw.Draw(f1)
    d.line([(120 + i * 3.2, 380 + 130 * math.sin(i / 9.0) * math.exp(-i / 220.0))
            for i in range(280)], fill=(20, 20, 25), width=6)
    d.line([(150, 470), (820, 430)], fill=(20, 20, 25), width=5)
    d.ellipse((828, 420, 844, 436), fill=(20, 20, 25))
    f1.save(firmas_dir / "firma_escaneo.png")
    f2 = Image.new("RGB", (1100, 700), (235, 228, 215))
    g = np.asarray(f2).astype(np.float32)
    yy, xx = np.mgrid[0:700, 0:1100]
    g *= (1.0 - 0.45 * (xx / 1100) * (yy / 700))[..., None]
    f2 = Image.fromarray(np.clip(g, 0, 255).astype(np.uint8))
    d = ImageDraw.Draw(f2)
    d.line([(140 + i * 3.0, 340 + 110 * math.sin(i / 7.5) * math.exp(-i / 260.0))
            for i in range(300)], fill=(35, 40, 80), width=7)
    ruido = np.random.default_rng(7).normal(0, 4, (700, 1100, 3))
    f2 = Image.fromarray(np.clip(np.asarray(f2).astype(np.float32) + ruido, 0, 255).astype(np.uint8))
    f2.save(firmas_dir / "firma_celular.png")

    trabajos = []  # (nombre_dorado, generador)
    for foto in entrada:
        for modo, transparente in (("transp", True), ("blanco", False)):
            trabajos.append((f"{foto.stem}_{modo}.png", ("foto", foto, transparente)))
    for firma in sorted(firmas_dir.iterdir()):
        trabajos.append((f"{firma.stem}.png", ("firma", firma, None)))

    sal = TMP / "salidas"
    sal.mkdir()
    core.SALIDA = sal
    n_ok = 0
    for nombre_dorado, (tipo, ruta, transparente) in trabajos:
        try:
            if tipo == "foto":
                preset["fondo_transparente"] = transparente
                destino, _, _ = core.procesar_una(ruta, preset, session,
                                                  Path(nombre_dorado).stem, fino=fino)
            else:
                destino = core.procesar_firma(ruta, Path(nombre_dorado).stem)
        except Exception as e:
            check(f"dorado {nombre_dorado}", False, f"reviento: {e}")
            continue
        dorada = DORADAS / destino.name
        if APROBAR:
            shutil.copy2(destino, dorada)
            n_ok += 1
        elif not dorada.exists():
            check(f"dorado {destino.name}", False,
                  "no existe la salida aprobada (corre --aprobar)")
        else:
            igual, det = comparar_imagenes(destino, dorada)
            check(f"dorado {destino.name}", igual, det)
            if igual:
                n_ok += 1
    if APROBAR:
        print(f"  APROBADAS {n_ok} salidas doradas en tests/doradas/ - "
              "revisalas a ojo antes de confiar en ellas.")

    # ---------- 6. humo de la interfaz ----------
    print("\n[6/6] Interfaz")
    try:
        import tkinter as tk
        import app
        root = tk.Tk()
        root.withdraw()
        a = app.App(root)
        root.update()
        tabs = [a.nb.tab(i, "text") for i in a.nb.tabs()]
        check("ventana arma con 4 pestañas", len(tabs) == 4, str(tabs))
        check("pestaña Renombrar/CardPresso (nueva)",
              any("CardPresso" in x for x in tabs)
              and hasattr(a, "renombrar_cardpresso"), str(tabs))
        check("boton Foto dificil existe", hasattr(a, "btn_dificil"))
        check("resolver confirmaciones disponible (Fase 2)",
              hasattr(a, "_resolver_confirmaciones") and hasattr(a, "resoluciones"))
        check("panel de resumen disponible (N4)",
              hasattr(a, "_mostrar_resumen_revision") and hasattr(a, "var_resumen"))
        check("boton procesar revisadas (C1)",
              hasattr(a, "btn_procesar_rev") and hasattr(a, "procesar_revisadas"))
        a._activar_botones(False)
        a._activar_botones(True)
        check("bloqueo de botones funciona", True)
        root.destroy()
    except Exception as e:
        check("interfaz arma", False, str(e)[:200])

    print(f"\n{'='*60}")
    if fallas:
        print(f"  RESULTADO: {len(fallas)} FALLA(S) en {time.time()-t0:.0f}s")
        for f in fallas:
            print(f"   - {f}")
    else:
        print(f"  RESULTADO: TODO OK ({time.time()-t0:.0f}s)")
    print("=" * 60)


if __name__ == "__main__":
    main()
    sys.exit(1 if fallas else 0)
